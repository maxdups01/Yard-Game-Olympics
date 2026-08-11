"""
Yard Game Olympics Scheduler
=============================
Generates a round-by-round schedule for a multi-team, multi-game "field day"
style tournament, and lays it out as a grid (rounds = rows, games = columns).

Requires: pandas, openpyxl (pip install pandas openpyxl)

Rules enforced:
  1. Every team plays every other team at least once (round robin).
  2. Every team plays every game at least once.
  3. Every team plays the same total number of matches (fair for scoring).

How it works:
  - A standard "circle method" round robin generates the round-by-round
    pairings (this already guarantees rule #1 and, as a side effect, rule #3 --
    every team plays exactly once per round, or sits out the same number of
    byes as everyone else if the team count is odd).
  - Games are then assigned to each match. A greedy algorithm tries, for every
    match, to hand both teams a game neither has played yet, breaking ties by
    (a) which game the two specific teams have repeated the least, then
    (b) which game is least used tournament-wide, then (c) randomly. This
    keeps not just the tournament-wide game totals even, but each team's own
    mix of games even too. It's retried with a number of random shuffles and
    the best (most complete) result is kept.
  - If the round robin doesn't produce enough matches for every team to see
    every game (e.g. lots of games, few teams), extra "makeup rounds" are
    appended -- one full extra round at a time, so every team still gets the
    same number of matches -- until every team has played every game.
  - Everything (schedule grid, team-vs-game counts, team-vs-team matchup
    counts) is written to one Excel workbook with three sheets, so it's easy
    to skim or hand-edit.
"""

import random
import itertools
import pandas as pd


# ---------------------------------------------------------------------------
# 1. Round robin pairing generator (circle method)
# ---------------------------------------------------------------------------
def round_robin_rounds(teams):
    """
    Returns a list of rounds; each round is a list of (teamA, teamB) tuples.
    Handles an odd number of teams by adding a BYE slot (that match is simply
    dropped from that round, so the team with the bye sits out).
    """
    arr = list(teams)
    if len(arr) % 2 == 1:
        arr.append("BYE")

    n = len(arr)
    num_rounds = n - 1
    rounds = []

    for _ in range(num_rounds):
        pairs = []
        for i in range(n // 2):
            t1, t2 = arr[i], arr[n - 1 - i]
            if t1 != "BYE" and t2 != "BYE":
                pairs.append((t1, t2))
        rounds.append(pairs)
        # rotate everyone except the first team
        arr = [arr[0]] + [arr[-1]] + arr[1:-1]

    return rounds


# ---------------------------------------------------------------------------
# 2. Greedy game assignment for a fixed set of rounds
# ---------------------------------------------------------------------------
def assign_round(matches, games, team_game_count, game_count, rnd):
    """
    Assigns a game to every match *within one round*, treating the round as
    a whole instead of match-by-match. This avoids double-booking a game
    (two matches playing the same game at the same time) whenever there are
    at least as many games as matches in the round. If there are genuinely
    more simultaneous matches than games (e.g. 8 games but only 5 stations
    worth of games for 12 matches), it spreads the repeats as evenly as
    possible instead of piling them onto one game.

    Tie-breaking priority when multiple games are equally "new" for a match:
      1. Prefer the game these two specific teams have (combined) played the
         fewest times so far -- this is what keeps each team's own game
         distribution even, not just the tournament-wide total.
      2. Fall back to the game used least often tournament-wide.
      3. Random, to avoid any bias from team/game ordering.

    Mutates `team_game_count` and `game_count` in place. Returns the round's
    match list as (teamA, teamB, game) tuples.
    """
    num_matches = len(matches)
    num_games = len(games)
    # max times any single game may be used in this round
    cap = -(-num_matches // num_games) if num_games else num_matches  # ceil division

    candidates = []
    for mi, (t1, t2) in enumerate(matches):
        c1 = team_game_count.setdefault(t1, {g: 0 for g in games})
        c2 = team_game_count.setdefault(t2, {g: 0 for g in games})
        for g in games:
            score = (c1[g] == 0) + (c2[g] == 0)          # 0/1/2: prefer unplayed
            pair_repeats = c1[g] + c2[g]                  # tie-break 1: per-team evenness
            candidates.append((score, pair_repeats, game_count[g], mi, g))

    rnd.shuffle(candidates)  # tie-break 3: random, before the stable sort
    candidates.sort(key=lambda c: (-c[0], c[1], c[2]))

    assigned_game = {}
    uses_this_round = {g: 0 for g in games}
    for score, _pair_repeats, _gc, mi, g in candidates:
        if mi in assigned_game or uses_this_round[g] >= cap:
            continue
        assigned_game[mi] = g
        uses_this_round[g] += 1
        if len(assigned_game) == num_matches:
            break

    round_out = []
    for mi, (t1, t2) in enumerate(matches):
        g = assigned_game[mi]
        team_game_count[t1][g] += 1
        team_game_count[t2][g] += 1
        game_count[g] += 1
        round_out.append((t1, t2, g))

    return round_out


def assign_games(rounds, games, seed):
    """
    Attempts to assign games to every round so that every team plays every
    game at least once, no game is double-booked within a round unless there
    are more simultaneous matches than games, and any leftover repeat games
    are spread evenly per team. Returns
    (schedule, team_game_count, game_count). schedule = list of rounds, each
    a list of (teamA, teamB, game) tuples. team_game_count = {team: {game: n}}.
    """
    rnd = random.Random(seed)
    team_game_count = {}          # team -> {game: times played}
    game_count = {g: 0 for g in games}  # game -> total times used, tournament-wide
    schedule = []

    for rnd_matches in rounds:
        matches = list(rnd_matches)
        round_out = assign_round(matches, games, team_game_count, game_count, rnd)
        schedule.append(round_out)

    return schedule, team_game_count, game_count


def _balance_score(team_game_count, teams, games):
    """Lower is better. Sum, over all teams, of (most-played game count minus
    least-played game count) for that team -- i.e. how spread out each
    team's own game mix is."""
    total = 0
    for t in teams:
        counts = [team_game_count.get(t, {}).get(g, 0) for g in games]
        total += max(counts) - min(counts)
    return total


def best_game_assignment(rounds, games, teams, attempts=300):
    """
    Try several random seeds. Prefers, in order: (1) fewest missing games
    across all teams, (2) the most even per-team game distribution. Stops
    early only if it finds a perfect result on both counts.
    """
    best = None
    best_key = (float("inf"), float("inf"))  # (missing_total, balance_score)

    for seed in range(attempts):
        schedule, team_game_count, game_count = assign_games(rounds, games, seed)
        missing_total = sum(
            sum(1 for g in games if team_game_count.get(t, {}).get(g, 0) == 0)
            for t in teams
        )
        balance_score = _balance_score(team_game_count, teams, games)
        key = (missing_total, balance_score)

        if key < best_key:
            best = (schedule, team_game_count, game_count)
            best_key = key
        if best_key == (0, 0) or (best_key[0] == 0 and best_key[1] <= 1):
            break

    assert best is not None  # attempts >= 1 guarantees this
    return best  # (schedule, team_game_count, game_count)


# ---------------------------------------------------------------------------
# 3. Makeup rounds -- add full extra rounds until every team has every game
# ---------------------------------------------------------------------------
def add_makeup_rounds(schedule, team_game_count, game_count, teams, games, max_rounds=200):
    """
    Appends complete extra rounds (every team plays, same as a normal round)
    until every team has played every game, or max_rounds safety cap is hit.
    Uses the same circle-method rotation so pairings stay varied and fair.
    """
    rr = round_robin_rounds(teams)  # reuse rotation pattern for makeup pairings
    rr_cycle = itertools.cycle(rr)
    rnd = random.Random(0)
    guard = 0

    def still_missing():
        return any(
            sum(1 for g in games if team_game_count.get(t, {}).get(g, 0) == 0) > 0
            for t in teams
        )

    while still_missing() and guard < max_rounds:
        pairs = next(rr_cycle)
        round_out = assign_round(pairs, games, team_game_count, game_count, rnd)
        schedule.append(round_out)
        guard += 1

    return schedule


# ---------------------------------------------------------------------------
# 4. Build the output grid (rounds = rows, games = columns)
# ---------------------------------------------------------------------------
def build_grid(schedule, games):
    rows = []
    for i, round_matches in enumerate(schedule, start=1):
        row: dict[str, object] = {"Round": i}
        cell_matches = {g: [] for g in games}
        for t1, t2, g in round_matches:
            cell_matches[g].append(f"{t1} vs {t2}")
        for g in games:
            row[g] = "; ".join(cell_matches[g]) if cell_matches[g] else ""
        rows.append(row)
    return pd.DataFrame(rows).set_index("Round")


# ---------------------------------------------------------------------------
# 5. Validation / summary
# ---------------------------------------------------------------------------
def team_game_counts(schedule, teams, games):
    """
    Returns a teams x games DataFrame: how many times each team played each
    game. Every cell should be >= 1 (rule #2), and ideally fairly even.
    """
    counts = {t: {g: 0 for g in games} for t in teams}
    for round_matches in schedule:
        for t1, t2, g in round_matches:
            counts[t1][g] += 1
            counts[t2][g] += 1
    return pd.DataFrame(counts).T[games]  # rows=teams, cols=games, in original game order


def matchup_counts(schedule, teams):
    """
    Returns a teams x teams DataFrame: how many times each pair of teams has
    played each other, regardless of which game it was. Diagonal is left as 0
    (a team doesn't play itself). Every off-diagonal cell should be >= 1.
    """
    counts = {t1: {t2: 0 for t2 in teams} for t1 in teams}
    for round_matches in schedule:
        for t1, t2, _g in round_matches:
            counts[t1][t2] += 1
            counts[t2][t1] += 1
    return pd.DataFrame(counts).T[teams]  # rows=teams, cols=teams, in original team order


def validate(schedule, teams, games):
    played = {t: set() for t in teams}
    matchups = set()
    match_counts = {t: 0 for t in teams}

    for round_matches in schedule:
        for t1, t2, g in round_matches:
            played[t1].add(g)
            played[t2].add(g)
            matchups.add(frozenset((t1, t2)))
            match_counts[t1] += 1
            match_counts[t2] += 1

    all_pairs = set(frozenset(p) for p in itertools.combinations(teams, 2))
    missing_matchups = all_pairs - matchups
    missing_games = {t: set(games) - played[t] for t in teams if set(games) - played[t]}

    double_booked = []  # (round_number, game, count) where a game is used >1x in a round
    max_matches_in_a_round = max((len(r) for r in schedule), default=0)
    for i, round_matches in enumerate(schedule, start=1):
        counts = {}
        for t1, t2, g in round_matches:
            counts[g] = counts.get(g, 0) + 1
        for g, c in counts.items():
            if c > 1:
                double_booked.append((i, g, c))

    print("\n--- Validation ---")
    print(f"Total rounds: {len(schedule)}")
    print(f"Matches per team: {match_counts}")
    if len(set(match_counts.values())) == 1:
        print("✔ Every team plays an equal number of matches.")
    else:
        print("⚠ Match counts are NOT perfectly equal (see above) -- "
              "this can happen with an odd number of teams.")

    if not missing_matchups:
        print("✔ Every team has played every other team at least once.")
    else:
        print(f"⚠ Missing matchups: {[tuple(m) for m in missing_matchups]}")

    if not missing_games:
        print("✔ Every team has played every game at least once.")
    else:
        print(f"⚠ Teams missing games: {missing_games}")

    if not double_booked:
        print("✔ No game is double-booked within a round.")
    elif len(games) >= max_matches_in_a_round:
        print(f"⚠ Unexpected double-booking (should have been avoidable): {double_booked}")
    else:
        print(f"ℹ Some rounds double-book a game (unavoidable -- more simultaneous "
              f"matches than games, so some games need 2 stations): {double_booked}")

    return match_counts, missing_matchups, missing_games, double_booked


# ---------------------------------------------------------------------------
# 6. Excel output -- one workbook, three sheets
# ---------------------------------------------------------------------------
def save_excel(path, grid, tg_counts, mm_counts):
    """
    Writes everything into a single .xlsx workbook so it's easy to open,
    skim, and hand-edit if needed:
      - "Schedule"            the rounds x games grid
      - "Team-Game Counts"    how many times each team played each game
      - "Team Matchup Counts" how many times each pair of teams has met
    """
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        grid.to_excel(writer, sheet_name="Schedule")
        tg_counts.to_excel(writer, sheet_name="Team-Game Counts")
        mm_counts.to_excel(writer, sheet_name="Team Matchup Counts")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # bold header row + index column, freeze so they stay visible when scrolling
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for cell in ws["A"]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "B2"
            # auto-size columns roughly based on content width
            for col_cells in ws.columns:
                col_letter = get_column_letter(col_cells[0].column)
                width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
                ws.column_dimensions[col_letter].width = min(width + 2, 40)
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
            ws["A1"].alignment = Alignment(horizontal="left")


# ---------------------------------------------------------------------------
# 7. Interactive entry point
# ---------------------------------------------------------------------------
def get_list_input(prompt, default):
    raw = input(f"{prompt} [default: {', '.join(default)}]: ").strip()
    if not raw:
        return default
    return [x.strip() for x in raw.split(",") if x.strip()]


def get_team_input(default_count=10):
    """
    Asks for a number of teams (generating "Team 1", "Team 2", ... automatically)
    or, if the person prefers, a comma-separated list of custom team names.
    """
    raw = input(
        f"Enter number of teams, or custom team names comma separated "
        f"[default: {default_count} teams]: "
    ).strip()

    if not raw:
        return [f"Team {i}" for i in range(1, default_count + 1)]

    if raw.isdigit():
        n = int(raw)
        if n < 2:
            print("Need at least 2 teams -- using the default instead.")
            return [f"T{i}" for i in range(1, default_count + 1)]
        return [f"T{i}" for i in range(1, n + 1)]

    # not a plain number -> treat as custom comma-separated team names
    return [x.strip() for x in raw.split(",") if x.strip()]


def main():
    default_team_count = 10
    default_games = ["Paddle Smash", "Cornhole", "KanJam",
                      "Bucket Golf", "Axe Throwing", "Texas Horseshoe"]

    print("=== Yard Game Olympics Scheduler ===")
    teams = get_team_input(default_team_count)
    games = get_list_input("Enter game names, comma separated", default_games)

    rounds = round_robin_rounds(teams)
    schedule, team_game_count, game_count = best_game_assignment(rounds, games, teams)

    still_missing = any(
        sum(1 for g in games if team_game_count.get(t, {}).get(g, 0) == 0) > 0
        for t in teams
    )
    if still_missing:
        print("\nBase round-robin schedule doesn't cover every game for every "
              "team -- adding makeup rounds...")
        schedule = add_makeup_rounds(schedule, team_game_count, game_count, teams, games)

    grid = build_grid(schedule, games)

    pd.set_option("display.width", 200)
    pd.set_option("display.max_columns", None)
    print("\n=== Schedule Grid (rows = rounds, columns = games) ===")
    print(grid)

    validate(schedule, teams, games)

    tg_counts = team_game_counts(schedule, teams, games)
    print("\n=== Times Each Team Played Each Game ===")
    print(tg_counts)

    mm_counts = matchup_counts(schedule, teams)
    print("\n=== Times Each Team Played Each Other Team (any game) ===")
    print(mm_counts)

    out_path = "yard_game_olympics_schedule.xlsx"
    save_excel(out_path, grid, tg_counts, mm_counts)
    print(f"\nSaved schedule + counts to {out_path} "
          f"(sheets: Schedule, Team-Game Counts, Team Matchup Counts)")


if __name__ == "__main__":
    main()
    